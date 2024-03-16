import openpyxl


def getRowCount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return sheet.max_row

def getColumnCount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return sheet.max_column

def get_cell_data(path,sheetname,rownum,columnnum):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return sheet.cell(row=rownum,column=columnnum).value

def set_cell_data(path,sheetname,rownum,columnnum,data):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    sheet.cell(row=rownum,column=columnnum).value=data
    workbook.save(path)

def get_data_from_excel(path, sheetname):

    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2,total_rows+1):
        row_list = []
        for c in range(1,total_cols+1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)

    return final_list
